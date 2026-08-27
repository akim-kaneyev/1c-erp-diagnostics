from __future__ import annotations

import io
import hashlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

try:
    from xlsx_profile import package_security_findings, safe_preview
except ModuleNotFoundError:  # pragma: no cover - importlib/package execution fallback
    from tools.xlsx_profile import package_security_findings, safe_preview


def safe_label(value: str, replacement: str) -> tuple[str, bool, bool]:
    preview, credential, machine_path = safe_preview(value)
    if credential or machine_path:
        return replacement, credential, machine_path
    return str(preview), False, False


def snapshot(path: Path) -> tuple[dict, bool, bool, bool, str]:
    payload = path.read_bytes()
    payload_hash = hashlib.sha256(payload).hexdigest()
    credential_exposure, machine_path_exposure, package_scan_incomplete = (
        package_security_findings(payload)
    )
    if credential_exposure or machine_path_exposure or package_scan_incomplete:
        # A failed package trust-boundary scan must not be followed by workbook
        # parsing; return only redacted control state and the stable file hash.
        return (
            {},
            credential_exposure,
            machine_path_exposure,
            package_scan_incomplete,
            payload_hash,
        )
    wb = None
    data = {}
    try:
        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
        for sheet_index, ws in enumerate(wb.worksheets, start=1):
            title, credential, machine_path = safe_label(
                ws.title, f"[REDACTED_SHEET_{sheet_index}]"
            )
            credential_exposure = credential_exposure or credential
            machine_path_exposure = machine_path_exposure or machine_path
            sheet = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value not in (None, ""):
                        value, credential, machine_path = safe_preview(cell.value)
                        credential_exposure = credential_exposure or credential
                        machine_path_exposure = machine_path_exposure or machine_path
                        sheet[cell.coordinate] = value
            data[title] = sheet
    except Exception:
        data = {}
        package_scan_incomplete = True
    finally:
        if wb is not None:
            wb.close()
    return (
        data,
        credential_exposure,
        machine_path_exposure,
        package_scan_incomplete,
        payload_hash,
    )


def main(a_name: str, b_name: str) -> int:
    try:
        a_path, b_path = Path(a_name), Path(b_name)
        a, a_credential, a_machine_path, a_incomplete, a_hash = snapshot(a_path)
        b, b_credential, b_machine_path, b_incomplete, b_hash = snapshot(b_path)
        all_sheets = sorted(set(a) | set(b))
        left_name, name_credential_a, name_machine_a = safe_label(
            a_path.name, "[REDACTED_LEFT_FILE]"
        )
        right_name, name_credential_b, name_machine_b = safe_label(
            b_path.name, "[REDACTED_RIGHT_FILE]"
        )
        credential_exposure = (
            a_credential or b_credential or name_credential_a or name_credential_b
        )
        machine_path_exposure = (
            a_machine_path or b_machine_path or name_machine_a or name_machine_b
        )
        package_scan_incomplete = a_incomplete or b_incomplete
        diff = {
            "status": "FAIL"
            if credential_exposure or machine_path_exposure or package_scan_incomplete
            else "PASS",
            "credential_exposure": credential_exposure,
            "machine_path_exposure": machine_path_exposure,
            "package_scan_incomplete": package_scan_incomplete,
            "left": left_name,
            "right": right_name,
            "sheets": {},
        }

        for s in all_sheets:
            sa, sb = a.get(s, {}), b.get(s, {})
            coords = sorted(set(sa) | set(sb))
            changes = []
            for c in coords:
                va, vb = sa.get(c), sb.get(c)
                if va != vb:
                    changes.append({"cell": c, "left": va, "right": vb})
            if changes:
                diff["sheets"][s] = changes

        left_stem, sensitive_left, _ = safe_label(a_path.stem, "redacted-left")
        right_stem, sensitive_right, _ = safe_label(b_path.stem, "redacted-right")
        if sensitive_left or sensitive_right:
            pair_hash = hashlib.sha256(
                f"{a_hash}\0{b_hash}".encode("ascii")
            ).hexdigest()[:16]
            out = Path(f"redacted-xlsx-comparison-{pair_hash}.diff.json")
        else:
            out = Path(f"{left_stem}__vs__{right_stem}.diff.json")
        out.write_text(
            json.dumps(diff, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
    except Exception:
        # Do not expose raw exception text: parsers and filesystem errors can embed
        # credential-like filenames, cell content and absolute machine paths.
        print("XLSX COMPARE: FAIL (input could not be processed safely)", file=sys.stderr)
        return 2
    print(f"Wrote {out.name}")
    return 0 if diff["status"] == "PASS" else 1


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python tools/compare_xlsx.py left.xlsx right.xlsx")
    raise SystemExit(main(sys.argv[1], sys.argv[2]))
