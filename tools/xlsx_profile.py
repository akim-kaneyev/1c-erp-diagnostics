from __future__ import annotations

import argparse
import codecs
import hashlib
import io
import json
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook


VIEW_KINDS = {"unknown", "rows", "property_tree"}
SCAN_TAIL_CHARS = 2048
MAX_PACKAGE_MEMBER_BYTES = 64 * 1024 * 1024
MAX_PACKAGE_TOTAL_BYTES = 256 * 1024 * 1024
MAX_PACKAGE_METADATA_BYTES = 64 * 1024
MAX_PACKAGE_METADATA_TOTAL_BYTES = 4 * 1024 * 1024
TEXT_PACKAGE_SUFFIXES = {".xml", ".rels", ".txt", ".csv", ".json", ".vml"}
EVIDENCE_ID = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.:-]{0,127}$")
SECRET_ASSIGNMENT = re.compile(
    r"(?i)(?:\\?[\"']?(?:password|passwd|api[_.-]?key|access[_.-]?token|refresh[_.-]?token|"
    r"client[_.-]?secret|private[_.-]?key|sonar[_.-]?token|"
    r"aws[_.-]?secret[_.-]?access[_.-]?key)\\?[\"']?|"
    r"sonar\.(?:token|login))\s*[:=]"
)
BEARER_AUTHORIZATION = re.compile(
    r"(?i)\\?[\"']?authorization\\?[\"']?\s*[:=]\s*\\?[\"']?\s*bearer\s+\S{8,}"
)
JSON_UNICODE_ESCAPE = re.compile(r"\\+[uU]([0-9A-Fa-f]{4})")
ESCAPED_QUOTE = re.compile(r"\\+([\"'])")
ABSOLUTE_MACHINE_PATH = re.compile(
    r"(?:file:/+(?:[A-Za-z]:[\\/]+Users[\\/]+[^\\/\s]+|"
    r"(?:Users|home)/[^/\s]+)|"
    r"(?<![A-Za-z0-9._:/-])[A-Za-z]:[\\/]+Users[\\/]+[^\\/\s]+|"
    r"(?<![A-Za-z0-9._:/-])/(?:Users|home)/[^/\s]+)",
    re.IGNORECASE,
)
PRIVATE_KEY_MARKERS = (
    "BEGIN " + "PRIVATE KEY",
    "BEGIN " + "RSA PRIVATE KEY",
    "BEGIN " + "EC PRIVATE KEY",
    "BEGIN " + "OPENSSH PRIVATE KEY",
    "BEGIN " + "ENCRYPTED PRIVATE KEY",
)
PRIVATE_KEY_HEADER = re.compile(
    r"-----BEGIN [^-\r\n]{0,80}" + "PRIVATE KEY" + r"(?: BLOCK)?-----",
    re.IGNORECASE,
)
PUTTY_PRIVATE_KEY_HEADER = re.compile(
    r"^\s*PuTTY-User-Key-File-[23]\s*:", re.IGNORECASE | re.MULTILINE
)
SSH2_PRIVATE_KEY_HEADER = re.compile(
    r"-{4,}\s*BEGIN\s+SSH2(?:\s+ENCRYPTED)?\s+PRIVATE\s+KEY\s*-{4,}",
    re.IGNORECASE,
)


def normalize_security_text(value: str) -> tuple[str, bool]:
    """Decode bounded JSON escaping; unresolved deep nesting fails closed."""
    normalized = value
    for _ in range(8):
        updated = JSON_UNICODE_ESCAPE.sub(
            lambda match: chr(int(match.group(1), 16)),
            normalized,
        )
        updated = ESCAPED_QUOTE.sub(lambda match: match.group(1), updated)
        if updated == normalized:
            break
        normalized = updated
    unresolved = JSON_UNICODE_ESCAPE.search(normalized) is not None
    return normalized, unresolved


def safe_preview(value):
    if not isinstance(value, str):
        return value, False, False
    normalized, unresolved_escape = normalize_security_text(value)
    upper = normalized.upper()
    credential = (
        unresolved_escape
        or SECRET_ASSIGNMENT.search(normalized) is not None
        or BEARER_AUTHORIZATION.search(normalized) is not None
        or PRIVATE_KEY_HEADER.search(normalized) is not None
        or PUTTY_PRIVATE_KEY_HEADER.search(normalized) is not None
        or SSH2_PRIVATE_KEY_HEADER.search(normalized) is not None
        or any(marker in upper for marker in PRIVATE_KEY_MARKERS)
    )
    machine_path = (
        ABSOLUTE_MACHINE_PATH.search(value) is not None
        or ABSOLUTE_MACHINE_PATH.search(normalized) is not None
    )
    if credential or machine_path:
        return "[REDACTED_SENSITIVE_VALUE]", credential, machine_path
    return value, False, False


def stream_security_findings(
    stream, *, scan_all_byte_phases: bool = False
) -> tuple[bool, bool]:
    # ZIP metadata can prefix encoded values with arbitrary bytes. Its bounded
    # callers opt into every UTF-16/UTF-32 phase; declared text-member streams
    # use their natural byte boundary and retain the smaller decoder set.
    states = [
        [codecs.getincrementaldecoder(encoding)(errors="ignore"), "", phase]
        for encoding, width in (
            ("utf-8", 1),
            ("utf-16-le", 2),
            ("utf-16-be", 2),
            ("utf-32-le", 4),
            ("utf-32-be", 4),
        )
        for phase in range(width if scan_all_byte_phases else 1)
    ]
    credential = False
    machine_path = False
    while True:
        chunk = stream.read(1024 * 1024)
        if not chunk:
            break
        for state in states:
            decoder, tail, phase = state
            payload = chunk
            if phase:
                skipped = min(int(phase), len(payload))
                payload = payload[skipped:]
                state[2] = int(phase) - skipped
            text = str(tail) + decoder.decode(payload)
            _, found_credential, found_machine_path = safe_preview(text)
            credential = credential or found_credential
            machine_path = machine_path or found_machine_path
            state[1] = text[-SCAN_TAIL_CHARS:]
    for state in states:
        decoder, tail, _ = state
        text = str(tail) + decoder.decode(b"", final=True)
        _, found_credential, found_machine_path = safe_preview(text)
        credential = credential or found_credential
        machine_path = machine_path or found_machine_path
    return credential, machine_path


def package_metadata_findings(payload: bytes) -> tuple[bool, bool, bool]:
    """Scan bounded ZIP metadata; the raw value is never returned or logged."""
    if len(payload) > MAX_PACKAGE_METADATA_BYTES:
        return False, False, True
    credential, machine_path = stream_security_findings(
        io.BytesIO(payload), scan_all_byte_phases=True
    )
    return credential, machine_path, False


def valid_extra_fields(payload: bytes) -> bool:
    """Validate the little-endian header/length framing of ZIP extra fields."""
    offset = 0
    while offset < len(payload):
        if len(payload) - offset < 4:
            return False
        field_size = int.from_bytes(payload[offset + 2 : offset + 4], "little")
        offset += 4
        if field_size > len(payload) - offset:
            return False
        offset += field_size
    return True


def local_header_metadata_findings(
    payload: bytes, member: zipfile.ZipInfo, remaining_budget: int
) -> tuple[bool, bool, bool, int]:
    """Inspect the local ZIP name/extra fields without scanning file payload bytes."""
    offset = member.header_offset
    fixed_size = 30
    if (
        isinstance(offset, bool)
        or not isinstance(offset, int)
        or offset < 0
        or offset + fixed_size > len(payload)
        or payload[offset : offset + 4] != b"PK\x03\x04"
    ):
        return False, False, True, 0
    name_length = int.from_bytes(payload[offset + 26 : offset + 28], "little")
    extra_length = int.from_bytes(payload[offset + 28 : offset + 30], "little")
    name_start = offset + fixed_size
    extra_start = name_start + name_length
    metadata_end = extra_start + extra_length
    if metadata_end > len(payload):
        return False, False, True, 0
    metadata_size = name_length + extra_length
    if metadata_size > remaining_budget:
        return False, False, True, 0

    credential = False
    machine_path = False
    incomplete = False
    local_name = payload[name_start:extra_start]
    local_extra = payload[extra_start:metadata_end]
    for metadata in (local_name, local_extra):
        found_credential, found_machine_path, metadata_incomplete = (
            package_metadata_findings(metadata)
        )
        credential = credential or found_credential
        machine_path = machine_path or found_machine_path
        incomplete = incomplete or metadata_incomplete
    if local_extra and not valid_extra_fields(local_extra):
        incomplete = True
    return credential, machine_path, incomplete, metadata_size


def central_directory_metadata_findings(
    payload: bytes,
) -> tuple[bool, bool, bool]:
    """Structurally scan raw central-directory names, extras and comments."""
    eocd_signature = b"PK\x05\x06"
    central_signature = b"PK\x01\x02"
    eocd_size = 22
    search_start = max(0, len(payload) - eocd_size - 0xFFFF)
    eocd_offset = payload.rfind(eocd_signature, search_start)
    if eocd_offset < 0 or eocd_offset + eocd_size > len(payload):
        return False, False, True

    comment_length = int.from_bytes(
        payload[eocd_offset + 20 : eocd_offset + 22], "little"
    )
    if eocd_offset + eocd_size + comment_length != len(payload):
        return False, False, True
    archive_comment = payload[eocd_offset + eocd_size :]
    credential, machine_path, incomplete = package_metadata_findings(archive_comment)

    disk_number = int.from_bytes(payload[eocd_offset + 4 : eocd_offset + 6], "little")
    central_disk = int.from_bytes(payload[eocd_offset + 6 : eocd_offset + 8], "little")
    entries_on_disk = int.from_bytes(
        payload[eocd_offset + 8 : eocd_offset + 10], "little"
    )
    entries_total = int.from_bytes(
        payload[eocd_offset + 10 : eocd_offset + 12], "little"
    )
    central_size = int.from_bytes(
        payload[eocd_offset + 12 : eocd_offset + 16], "little"
    )
    central_offset = int.from_bytes(
        payload[eocd_offset + 16 : eocd_offset + 20], "little"
    )
    if (
        disk_number != 0
        or central_disk != 0
        or entries_on_disk != entries_total
        or entries_total == 0xFFFF
        or central_size == 0xFFFFFFFF
        or central_offset == 0xFFFFFFFF
        or central_size > MAX_PACKAGE_METADATA_TOTAL_BYTES
    ):
        return credential, machine_path, True

    # ZIP offsets are relative to the archive, which can be prefixed by a
    # self-extracting stub. Derive the same concatenation adjustment as readers.
    concatenated_prefix = eocd_offset - central_size - central_offset
    central_start = central_offset + concatenated_prefix
    central_end = central_start + central_size
    if central_start < 0 or central_end != eocd_offset:
        return credential, machine_path, True

    cursor = central_start
    for _ in range(entries_total):
        fixed_size = 46
        if (
            cursor < 0
            or cursor + fixed_size > central_end
            or payload[cursor : cursor + 4] != central_signature
        ):
            return credential, machine_path, True
        name_length = int.from_bytes(payload[cursor + 28 : cursor + 30], "little")
        extra_length = int.from_bytes(payload[cursor + 30 : cursor + 32], "little")
        entry_comment_length = int.from_bytes(
            payload[cursor + 32 : cursor + 34], "little"
        )
        name_start = cursor + fixed_size
        extra_start = name_start + name_length
        comment_start = extra_start + extra_length
        entry_end = comment_start + entry_comment_length
        if entry_end > central_end:
            return credential, machine_path, True
        central_extra = payload[extra_start:comment_start]
        for metadata in (
            payload[name_start:extra_start],
            central_extra,
            payload[comment_start:entry_end],
        ):
            found_credential, found_machine_path, metadata_incomplete = (
                package_metadata_findings(metadata)
            )
            credential = credential or found_credential
            machine_path = machine_path or found_machine_path
            incomplete = incomplete or metadata_incomplete
        if central_extra and not valid_extra_fields(central_extra):
            incomplete = True
        cursor = entry_end
    if cursor != central_end:
        incomplete = True
    return credential, machine_path, incomplete


def package_security_findings(payload: bytes) -> tuple[bool, bool, bool]:
    """Inspect XLSX package metadata without exposing matched values."""
    if len(payload) > MAX_PACKAGE_TOTAL_BYTES:
        return False, False, True
    # Central-directory metadata is not authoritative: a local ZIP header can
    # carry different extra bytes. Parse both header copies structurally and do
    # not decode the complete compressed package as text.
    credential, machine_path, incomplete = central_directory_metadata_findings(payload)
    if incomplete:
        return credential, machine_path, True
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            members = archive.infolist()
            local_metadata_budget = MAX_PACKAGE_METADATA_TOTAL_BYTES
            total_size_exceeded = (
                sum(member.file_size for member in members) > MAX_PACKAGE_TOTAL_BYTES
            )
            if total_size_exceeded:
                incomplete = True
            archive_credential, archive_machine_path, archive_incomplete = (
                package_metadata_findings(archive.comment)
            )
            credential = credential or archive_credential
            machine_path = machine_path or archive_machine_path
            incomplete = incomplete or archive_incomplete
            for member in members:
                (
                    local_credential,
                    local_machine_path,
                    local_incomplete,
                    local_metadata_size,
                ) = local_header_metadata_findings(
                    payload, member, local_metadata_budget
                )
                credential = credential or local_credential
                machine_path = machine_path or local_machine_path
                incomplete = incomplete or local_incomplete
                if local_incomplete:
                    return credential, machine_path, True
                local_metadata_budget -= local_metadata_size
                safe_name, name_credential, name_machine_path = safe_preview(member.filename)
                del safe_name
                credential = credential or name_credential
                machine_path = machine_path or name_machine_path
                for metadata in (member.comment, member.extra):
                    metadata_credential, metadata_machine_path, metadata_incomplete = (
                        package_metadata_findings(metadata)
                    )
                    credential = credential or metadata_credential
                    machine_path = machine_path or metadata_machine_path
                    incomplete = incomplete or metadata_incomplete
                if member.extra and not valid_extra_fields(member.extra):
                    incomplete = True
                normalized_name = member.filename.replace("\\", "/").lower()
                if member.is_dir() and member.file_size == 0 and member.compress_size == 0:
                    # Explicit ZIP directory entries carry no content. Their names
                    # were already checked above, so they are not unscanned binaries.
                    continue
                member_name = normalized_name.rsplit("/", 1)[-1]
                suffix = Path(member_name).suffix
                if not suffix and member_name.startswith("."):
                    # OPC root relationships are conventionally stored as
                    # `_rels/.rels`; pathlib treats that dotfile as suffixless.
                    suffix = member_name
                if (
                    total_size_exceeded
                    or member.flag_bits & 0x1
                    or member.file_size > MAX_PACKAGE_MEMBER_BYTES
                ):
                    incomplete = True
                    continue
                if member.compress_size and member.file_size > member.compress_size * 1000:
                    incomplete = True
                    continue
                if suffix not in TEXT_PACKAGE_SUFFIXES:
                    # A non-text package member cannot be proven free of credentials
                    # or machine-local data by the bounded text scanner. This applies
                    # to every location, including VBA projects and media, not only
                    # xl/embeddings. Standard styles are XML and remain scannable.
                    incomplete = True
                    continue
                try:
                    with archive.open(member) as stream:
                        found_credential, found_machine_path = stream_security_findings(stream)
                except (OSError, RuntimeError, zipfile.BadZipFile):
                    incomplete = True
                    continue
                credential = credential or found_credential
                machine_path = machine_path or found_machine_path
    except (OSError, RuntimeError, zipfile.BadZipFile):
        return credential, machine_path, True
    return credential, machine_path, incomplete


def inspect_workbook_rows(
    wb,
    result: dict,
    view_kind: str,
    view_sheet_index: int | None,
) -> None:
    """Populate a redacted profile; callers handle lazy parser failures."""
    for sheet_index, ws in enumerate(wb.worksheets, start=1):
        safe_title, title_credential, title_machine_path = safe_preview(ws.title)
        result["credential_exposure"] = result["credential_exposure"] or title_credential
        result["machine_path_exposure"] = (
            result["machine_path_exposure"] or title_machine_path
        )
        if title_credential or title_machine_path:
            safe_title = f"[REDACTED_SHEET_{sheet_index}]"
        nonempty = 0
        formulas = 0
        samples = []
        nonempty_rows = 0
        used_columns = 0
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            row_values = []
            row_nonempty = False
            for cell in row:
                value = cell.value
                if value not in (None, ""):
                    nonempty += 1
                    row_nonempty = True
                    if isinstance(value, str) and value.startswith("="):
                        formulas += 1
                preview_value, credential, machine_path = safe_preview(value)
                result["credential_exposure"] = result["credential_exposure"] or credential
                result["machine_path_exposure"] = (
                    result["machine_path_exposure"] or machine_path
                )
                row_values.append(preview_value)
            if row_nonempty and len(samples) < 10:
                samples.append({"row": row_idx, "values": row_values[:30]})
            if row_nonempty:
                nonempty_rows += 1
                used_columns = max(
                    used_columns,
                    max(
                        (
                            index
                            for index, item in enumerate(row_values, start=1)
                            if item not in (None, "")
                        ),
                        default=0,
                    ),
                )

        candidate_rows = nonempty_rows >= 2
        result["candidate_tabular_rows_present"] = (
            result["candidate_tabular_rows_present"] or candidate_rows
        )
        visual_gate_applies = sheet_index == view_sheet_index
        rows_present = visual_gate_applies and view_kind == "rows" and candidate_rows
        result["rows_present"] = result["rows_present"] or rows_present

        result["sheets"].append(
            {
                "title": safe_title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "row_count": nonempty_rows,
                "column_count": used_columns,
                "candidate_tabular_rows_present": candidate_rows,
                "visual_gate_applies": visual_gate_applies,
                "rows_present": rows_present,
                "view_status": (
                    "rows_present"
                    if rows_present
                    else "wrong_view/incomplete"
                    if visual_gate_applies and view_kind in {"rows", "property_tree"}
                    else "unverified/incomplete"
                ),
                "header_preview": samples[0]["values"] if samples else [],
                "value_preview": [item["values"] for item in samples[1:4]],
                "nonempty_cells": nonempty,
                "formula_cells": formulas,
                "sample_nonempty_rows": samples,
            }
        )


def profile(
    path: Path,
    *,
    view_kind: str = "unknown",
    view_evidence_id: str | None = None,
    view_sheet_index: int | None = None,
) -> dict:
    if view_kind not in VIEW_KINDS:
        raise ValueError("view_kind must be unknown, rows or property_tree")
    normalized_view_evidence_id = (
        view_evidence_id.strip() if isinstance(view_evidence_id, str) else None
    )
    if normalized_view_evidence_id is not None and not EVIDENCE_ID.fullmatch(
        normalized_view_evidence_id
    ):
        raise ValueError("view_evidence_id must be a safe Evidence identifier")
    if normalized_view_evidence_id is not None:
        _, evidence_credential, evidence_machine_path = safe_preview(
            normalized_view_evidence_id
        )
        if evidence_credential or evidence_machine_path:
            raise ValueError("view_evidence_id must be a safe Evidence identifier")
    if view_kind != "unknown" and not (
        normalized_view_evidence_id
    ):
        raise ValueError("confirmed view_kind requires a visual-gate Evidence ID")
    if view_kind != "unknown" and (
        isinstance(view_sheet_index, bool)
        or not isinstance(view_sheet_index, int)
        or view_sheet_index < 1
    ):
        raise ValueError("confirmed view_kind requires a positive view_sheet_index")
    if view_kind == "unknown" and view_sheet_index is not None:
        raise ValueError("view_sheet_index requires a confirmed view_kind")
    payload = path.read_bytes()
    package_credential, package_machine_path, package_scan_incomplete = (
        package_security_findings(payload)
    )
    safe_file_name, file_credential, file_machine_path = safe_preview(path.name)
    result = {
        "file": safe_file_name,
        "sha256": hashlib.sha256(payload).hexdigest(),
        "status": "PASS",
        "view_kind": view_kind,
        "view_evidence_id": normalized_view_evidence_id,
        "view_sheet_index": view_sheet_index,
        "candidate_tabular_rows_present": False,
        "rows_present": False,
        "view_status": "unverified/incomplete",
        "credential_exposure": file_credential or package_credential,
        "machine_path_exposure": file_machine_path or package_machine_path,
        "package_scan_incomplete": package_scan_incomplete,
        "sheets": [],
    }
    if (
        result["credential_exposure"]
        or result["machine_path_exposure"]
        or result["package_scan_incomplete"]
    ):
        # Do not pass a package that already failed its trust-boundary scan to
        # the workbook parser. This also bounds parser work for oversized or
        # structurally unsupported archives.
        result["status"] = "FAIL"
        result["view_status"] = (
            "sensitive_content/incomplete"
            if result["credential_exposure"] or result["machine_path_exposure"]
            else "package_scan_incomplete"
        )
        return result
    try:
        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    except Exception:
        # Parser failures are a trust-boundary result, never a traceback containing
        # workbook metadata. Package findings above remain redacted and preserved.
        result["status"] = "FAIL"
        result["rows_present"] = False
        result["package_scan_incomplete"] = True
        result["view_status"] = (
            "sensitive_content/incomplete"
            if result["credential_exposure"] or result["machine_path_exposure"]
            else "package_scan_incomplete"
        )
        return result
    if view_sheet_index is not None and view_sheet_index > len(wb.worksheets):
        wb.close()
        raise ValueError("view_sheet_index is outside the workbook sheet range")

    try:
        inspect_workbook_rows(wb, result, view_kind, view_sheet_index)
    except Exception:
        result["status"] = "FAIL"
        result["rows_present"] = False
        result["package_scan_incomplete"] = True
        result["view_status"] = (
            "sensitive_content/incomplete"
            if result["credential_exposure"] or result["machine_path_exposure"]
            else "package_scan_incomplete"
        )
        result["sheets"] = []
        return result
    finally:
        wb.close()
    if (
        result["credential_exposure"]
        or result["machine_path_exposure"]
        or result["package_scan_incomplete"]
    ):
        result["status"] = "FAIL"
        result["rows_present"] = False
        result["view_status"] = (
            "sensitive_content/incomplete"
            if result["credential_exposure"] or result["machine_path_exposure"]
            else "package_scan_incomplete"
        )
        for sheet in result["sheets"]:
            sheet["rows_present"] = False
            sheet["view_status"] = result["view_status"]
    elif result["rows_present"]:
        result["view_status"] = "rows_present"
    elif view_kind in {"rows", "property_tree"}:
        result["view_status"] = "wrong_view/incomplete"
    return result


class RedactingArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> None:
        del message
        self.exit(2, "xlsx_profile.py: error: invalid arguments\n")


def build_parser() -> argparse.ArgumentParser:
    parser = RedactingArgumentParser(
        prog="xlsx_profile.py",
        description="Create a redacted XLSX evidence profile",
    )
    parser.add_argument("input", type=Path)
    parser.add_argument("--view-kind", default="unknown")
    parser.add_argument("--view-evidence-id")
    parser.add_argument("--view-sheet-index")
    return parser


def main(argv=None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    if args.view_kind != "unknown" and not (
        isinstance(args.view_evidence_id, str) and args.view_evidence_id.strip()
    ):
        parser.error("confirmed view kind requires evidence")
    view_sheet_index = None
    if args.view_sheet_index is not None:
        try:
            view_sheet_index = int(args.view_sheet_index)
        except (TypeError, ValueError):
            parser.error("sheet index must be an integer")
    if args.view_kind != "unknown" and (
        not isinstance(view_sheet_index, int) or view_sheet_index < 1
    ):
        parser.error("confirmed view kind requires a sheet index")
    path = args.input
    try:
        data = profile(
            path,
            view_kind=args.view_kind,
            view_evidence_id=args.view_evidence_id,
            view_sheet_index=view_sheet_index,
        )
        _, file_credential, file_machine_path = safe_preview(path.name)
        out = (
            path.parent / f"redacted-xlsx-profile-{data['sha256'][:16]}.json"
            if file_credential or file_machine_path
            else path.with_suffix(path.suffix + ".profile.json")
        )
        out.write_text(
            json.dumps(data, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
    except Exception:
        # Exception messages may contain an input filename, an absolute machine path,
        # workbook cell text or credentials. Keep the CLI boundary fail-closed and
        # never echo the raw exception. Programmatic callers can use profile().
        print("XLSX PROFILE: FAIL (input could not be processed safely)", file=sys.stderr)
        return 2
    print(f"Wrote {out.name}")
    return 0 if data["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
