from __future__ import annotations

import hashlib
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from pypdf import PdfWriter

ROOT = Path(__file__).resolve().parents[1]


class ToolRegressionTests(unittest.TestCase):
    def run_tool(self, tool: str, *args: str, cwd: Path | None = None) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(ROOT / "tools" / tool), *args],
            cwd=cwd or ROOT,
            check=True,
            capture_output=True,
            text=True,
        )

    def test_index_case_creates_hash_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            case = Path(tmp) / "case"
            input_dir = case / "input"
            input_dir.mkdir(parents=True)
            source = input_dir / "sample.txt"
            source.write_text("evidence", encoding="utf-8")

            self.run_tool("index_case.py", str(case))

            manifest = json.loads((case / "manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest["case"], "case")
            self.assertEqual(len(manifest["files"]), 1)
            self.assertEqual(
                manifest["files"][0]["sha256"],
                hashlib.sha256(b"evidence").hexdigest(),
            )

    def test_xlsx_profile_and_compare(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            left = work / "left.xlsx"
            right = work / "right.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws["A1"] = "Item"
            ws["B1"] = 10
            wb.save(left)

            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws["A1"] = "Item"
            ws["B1"] = 20
            wb.save(right)

            self.run_tool("xlsx_profile.py", str(left), cwd=work)
            profile = json.loads((work / "left.xlsx.profile.json").read_text(encoding="utf-8"))
            self.assertEqual(profile["sheets"][0]["title"], "Data")
            self.assertFalse(profile["rows_present"])
            self.assertEqual(profile["view_status"], "unverified/incomplete")
            self.assertEqual(profile["sheets"][0]["row_count"], 1)
            self.assertEqual(profile["sheets"][0]["column_count"], 2)
            self.assertEqual(profile["sheets"][0]["header_preview"], ["Item", 10])
            self.assertEqual(profile["sheets"][0]["value_preview"], [])
            self.assertRegex(profile["sha256"], r"^[0-9a-f]{64}$")

            self.run_tool("compare_xlsx.py", str(left), str(right), cwd=work)
            diff = json.loads((work / "left__vs__right.diff.json").read_text(encoding="utf-8"))
            self.assertEqual(diff["sheets"]["Data"][0]["cell"], "B1")
            self.assertEqual(diff["sheets"]["Data"][0]["left"], 10)
            self.assertEqual(diff["sheets"]["Data"][0]["right"], 20)

    def test_xlsx_property_tree_requires_explicit_visual_classification(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            artifact = work / "property-tree.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Property", "Name"])
            ws.append(["Type", "String"])
            wb.save(artifact)

            self.run_tool(
                "xlsx_profile.py",
                str(artifact),
                "--view-kind",
                "property_tree",
                "--view-evidence-id",
                "E-VIEW-1",
                "--view-sheet-index",
                "1",
                cwd=work,
            )
            profile = json.loads(
                (work / "property-tree.xlsx.profile.json").read_text(encoding="utf-8")
            )
            self.assertTrue(profile["candidate_tabular_rows_present"])
            self.assertFalse(profile["rows_present"])
            self.assertEqual(profile["view_status"], "wrong_view/incomplete")
            self.assertEqual(profile["view_evidence_id"], "E-VIEW-1")
            self.assertEqual(profile["view_sheet_index"], 1)

    def test_pdf_extract_handles_blank_page(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            pdf = work / "blank.pdf"
            writer = PdfWriter()
            writer.add_blank_page(width=300, height=300)
            with pdf.open("wb") as stream:
                writer.write(stream)

            self.run_tool("pdf_extract.py", str(pdf), cwd=work)
            text = (work / "blank.pdf.txt").read_text(encoding="utf-8")
            self.assertIn("===== PAGE 1 =====", text)


if __name__ == "__main__":
    unittest.main()
