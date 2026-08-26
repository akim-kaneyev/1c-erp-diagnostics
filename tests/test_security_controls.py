from __future__ import annotations

import importlib.util
import io
import json
import os
import subprocess
import sys
import tarfile
import tempfile
import unittest
import zipfile
from contextlib import redirect_stderr, redirect_stdout
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl.comments import Comment
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]


def load_module(name: str, relative: str):
    path = ROOT / relative
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


HISTORY = load_module("validate_publication_history", "tools/validate_publication_history.py")
PUBLIC = load_module("validate_public_release", "tools/validate_public_release.py")
XLSX = load_module("security_xlsx_profile", "tools/xlsx_profile.py")
COMPARE_XLSX = load_module("security_compare_xlsx", "tools/compare_xlsx.py")
CASE_STATE = load_module(
    "security_validate_case_state",
    "plugins/one-c-erp-diagnostics/skills/one-c-erp-case-state/scripts/validate_case_state.py",
)


def tar_with_bytes(path: str, data: bytes) -> bytes:
    buffer = io.BytesIO()
    with tarfile.open(fileobj=buffer, mode="w") as archive:
        info = tarfile.TarInfo(path)
        info.size = len(data)
        archive.addfile(info, io.BytesIO(data))
    return buffer.getvalue()


def tar_with_text(path: str, payload: str) -> bytes:
    return tar_with_bytes(path, payload.encode("utf-8"))


def tar_with_symlink(path: str, target: str) -> bytes:
    buffer = io.BytesIO()
    with tarfile.open(fileobj=buffer, mode="w") as archive:
        info = tarfile.TarInfo(path)
        info.type = tarfile.SYMTYPE
        info.linkname = target
        archive.addfile(info)
    return buffer.getvalue()


def zip_with_text(path: str, payload: str) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(path, payload)
    return buffer.getvalue()


class SecurityControlTests(unittest.TestCase):
    def test_pathless_blobs_and_commit_or_tag_metadata_are_scanned(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "reachable_object_1234"
        payload = f"{field}={marker}".encode("utf-8")
        object_types = {
            "a" * 40: "blob",
            "b" * 40: "commit",
            "c" * 40: "tag",
        }

        counts, errors, tree_ids = HISTORY.inspect_reachable_objects(
            object_types,
            lambda oid: object_types[oid],
            lambda _oid, _kind: HISTORY.stream_text_findings(io.BytesIO(payload)),
        )

        self.assertEqual(counts, {"blob": 1, "commit": 1, "tag": 1})
        self.assertEqual(tree_ids, set())
        rendered = "\n".join(errors)
        self.assertEqual(rendered.count("plaintext credential"), 3)
        self.assertIn("PATHLESS_GIT_BLOB", rendered)
        self.assertIn("PATHLESS_GIT_COMMIT", rendered)
        self.assertIn("PATHLESS_GIT_TAG", rendered)
        self.assertNotIn(marker, rendered)

        _, path_errors, _ = HISTORY.inspect_reachable_objects(
            ["d" * 40 + " artifact.dt"],
            lambda _oid: "blob",
            lambda _oid, _kind: (False, False, False),
        )
        self.assertTrue(any("forbidden artifact suffix" in item for item in path_errors))

    def test_standalone_tagged_tree_cannot_hide_forbidden_history_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)

            def run_git(*args: str, payload: bytes | None = None) -> bytes:
                result = subprocess.run(
                    ["git", *args],
                    cwd=work,
                    input=payload,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    check=True,
                )
                return result.stdout

            run_git("init", "-q")
            run_git("config", "user.name", "Synthetic Test")
            run_git("config", "user.email", "synthetic@users.noreply.github.com")
            (work / "safe.txt").write_text("safe", encoding="utf-8")
            run_git("add", "safe.txt")
            run_git("commit", "-q", "-m", "safe")
            blob_id = run_git(
                "hash-object", "-w", "--stdin", payload=b"opaque synthetic artifact"
            ).decode("ascii").strip()
            tree_id = run_git(
                "mktree",
                payload=f"100644 blob {blob_id}\tartifact.dt\n".encode("ascii"),
            ).decode("ascii").strip()
            run_git("tag", "forbidden-tree", tree_id)

            result = subprocess.run(
                [sys.executable, str(ROOT / "tools" / "validate_publication_history.py")],
                cwd=work,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                check=False,
            )

        self.assertEqual(result.returncode, 1)
        rendered = result.stdout + result.stderr
        self.assertIn("forbidden artifact suffix", rendered)
        self.assertNotIn("opaque synthetic artifact", rendered)

    def test_nested_safe_cases_path_keeps_its_repository_prefix(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)

            def run_git(*args: str) -> subprocess.CompletedProcess[str]:
                return subprocess.run(
                    ["git", *args],
                    cwd=work,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    check=True,
                )

            run_git("init", "-q")
            run_git("config", "user.name", "Synthetic Test")
            run_git("config", "user.email", "synthetic@users.noreply.github.com")
            safe_dir = work / "evals" / "cases"
            safe_dir.mkdir(parents=True)
            (safe_dir / "synthetic.json").write_text("{}\n", encoding="utf-8")
            run_git("add", "evals/cases/synthetic.json")
            run_git("commit", "-q", "-m", "safe nested case")

            result = subprocess.run(
                [sys.executable, str(ROOT / "tools" / "validate_publication_history.py")],
                cwd=work,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                check=False,
            )

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertIn("PUBLICATION HISTORY VALIDATION: PASS", result.stdout)

    def test_standalone_tree_root_keeps_root_cases_policy(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)

            def run_git(*args: str, payload: bytes | None = None) -> bytes:
                result = subprocess.run(
                    ["git", *args],
                    cwd=work,
                    input=payload,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    check=True,
                )
                return result.stdout

            run_git("init", "-q")
            run_git("config", "user.name", "Synthetic Test")
            run_git("config", "user.email", "synthetic@users.noreply.github.com")
            (work / "safe.txt").write_text("safe", encoding="utf-8")
            run_git("add", "safe.txt")
            run_git("commit", "-q", "-m", "safe")
            blob_id = run_git(
                "hash-object", "-w", "--stdin", payload=b"sanitized synthetic case"
            ).decode("ascii").strip()
            cases_tree = run_git(
                "mktree",
                payload=f"100644 blob {blob_id}\tsynthetic.json\n".encode("ascii"),
            ).decode("ascii").strip()
            root_tree = run_git(
                "mktree",
                payload=f"040000 tree {cases_tree}\tcases\n".encode("ascii"),
            ).decode("ascii").strip()
            run_git("tag", "standalone-cases-tree", root_tree)

            result = subprocess.run(
                [sys.executable, str(ROOT / "tools" / "validate_publication_history.py")],
                cwd=work,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                check=False,
            )

        self.assertEqual(result.returncode, 1)
        rendered = result.stdout + result.stderr
        self.assertIn("case data", rendered)
        self.assertIn("cases/synthetic.json", rendered)

    def test_shared_blob_alias_in_tagged_tree_cannot_hide_forbidden_path(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)

            def run_git(*args: str, payload: bytes | None = None) -> bytes:
                result = subprocess.run(
                    ["git", *args],
                    cwd=work,
                    input=payload,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    check=True,
                )
                return result.stdout

            run_git("init", "-q")
            run_git("config", "user.name", "Synthetic Test")
            run_git("config", "user.email", "synthetic@users.noreply.github.com")
            (work / "base.txt").write_text("base", encoding="utf-8")
            run_git("add", "base.txt")
            run_git("commit", "-q", "-m", "base")
            blob_id = run_git(
                "hash-object", "-w", "--stdin", payload=b"shared opaque content"
            ).decode("ascii").strip()
            safe_tree_id = run_git(
                "mktree",
                payload=f"100644 blob {blob_id}\tsafe.txt\n".encode("ascii"),
            ).decode("ascii").strip()
            forbidden_tree_id = run_git(
                "mktree",
                payload=f"100644 blob {blob_id}\tartifact.dt\n".encode("ascii"),
            ).decode("ascii").strip()
            # rev-list reports the shared blob only once, under the first tree's
            # safe alias. The validator must still inspect both tree entries.
            run_git("tag", "a-safe", safe_tree_id)
            run_git("tag", "z-forbidden", forbidden_tree_id)

            result = subprocess.run(
                [sys.executable, str(ROOT / "tools" / "validate_publication_history.py")],
                cwd=work,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                check=False,
            )

        self.assertEqual(result.returncode, 1)
        rendered = result.stdout + result.stderr
        self.assertIn("forbidden artifact suffix", rendered)
        self.assertNotIn("shared opaque content", rendered)

    def test_standalone_tagged_tree_cannot_hide_unsafe_symlink(self) -> None:
        target = "../../synthetic-outside"
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)

            def run_git(*args: str, payload: bytes | None = None) -> bytes:
                result = subprocess.run(
                    ["git", *args],
                    cwd=work,
                    input=payload,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    check=True,
                )
                return result.stdout

            run_git("init", "-q")
            run_git("config", "user.name", "Synthetic Test")
            run_git("config", "user.email", "synthetic@users.noreply.github.com")
            (work / "safe.txt").write_text("safe", encoding="utf-8")
            run_git("add", "safe.txt")
            run_git("commit", "-q", "-m", "safe")
            target_blob = run_git(
                "hash-object", "-w", "--stdin", payload=target.encode("utf-8")
            ).decode("ascii").strip()
            tree_id = run_git(
                "mktree",
                payload=f"120000 blob {target_blob}\tsafe-link\n".encode("ascii"),
            ).decode("ascii").strip()
            run_git("tag", "unsafe-link-tree", tree_id)

            result = subprocess.run(
                [sys.executable, str(ROOT / "tools" / "validate_publication_history.py")],
                cwd=work,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                check=False,
            )

        self.assertEqual(result.returncode, 1)
        rendered = result.stdout + result.stderr
        self.assertIn("unsafe historical link target", rendered)
        self.assertNotIn(target, rendered)

    def test_tree_and_archive_secret_detection_never_echoes_value(self) -> None:
        field = "pass" + "word"
        value = "synthetic_" + "marker_" * 4
        payload = f"{field}={value}"

        self.assertIsNotNone(PUBLIC.SECRET_ASSIGNMENT.search(payload))
        archived, errors = HISTORY.inspect_archive(tar_with_text("report.txt", payload))

        self.assertEqual(archived, ["report.txt"])
        self.assertTrue(any("credential" in item and "report.txt" in item for item in errors))
        self.assertNotIn(value, "\n".join(errors))

        state_result = CASE_STATE.validate_state(
            {
                "schema_version": 1,
                "case_id": payload,
                "evidence": [],
                "runs": [],
                "claims": [],
                "documents": [],
                "gates": [],
                "active_index": {},
            }
        )
        rendered = str(state_result)
        self.assertEqual(state_result["errors"][0]["code"], "credential_exposure")
        self.assertNotIn(value, rendered)

    def test_cache_named_directories_do_not_hide_arbitrary_sensitive_files(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "cache_secret_1234"
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            cache = root / "docs" / "__pycache__"
            cache.mkdir(parents=True)
            hidden = cache / "credential.json"
            hidden.write_text(f"{field}={marker}", encoding="utf-8")
            self.assertFalse(PUBLIC.is_verified_generated_python_cache(hidden))
            secret, _ = PUBLIC.content_findings(hidden.read_bytes())
            self.assertTrue(secret)

            source = root / "module.py"
            source.write_text("value = 1\n", encoding="utf-8")
            generated_cache = root / "__pycache__"
            generated_cache.mkdir()
            pyc = generated_cache / "module.cpython-312.pyc"
            pyc.write_bytes(importlib.util.MAGIC_NUMBER + b"synthetic-generated-bytecode")
            self.assertTrue(PUBLIC.is_verified_generated_python_cache(pyc))

    def test_json_utf16_and_private_key_markers_are_redacted(self) -> None:
        field = "pass" + "word"
        value = "synthetic_" + "json_marker_1234"
        json_payload = "{\"" + field + "\":\"" + value + "\"}"
        nested_payload = "{\"" + field + "\":{\"value\":\"" + value + "\"}}"
        escaped_payload = json_payload.replace('"', '\\"')
        unicode_key_payload = '{"pass' + "\\u0077" + 'ord":"' + value + '"}'
        dotted_field = "api" + ".key"
        dotted_key_payload = '{"' + dotted_field + '":{"value":"' + value + '"}}'
        private_marker = "-----" + "BEGIN " + "OPENSSH PRIVATE KEY" + "-----"
        encrypted_marker = "-----" + "BEGIN " + "ENCRYPTED PRIVATE KEY" + "-----"
        dsa_marker = "-----" + "BEGIN DSA " + "PRIVATE KEY" + "-----"
        pgp_marker = "-----" + "BEGIN PGP " + "PRIVATE KEY" + " BLOCK-----"
        ppk_marker = (
            "PuTTY-User-Key-File-3: ssh-rsa\n"
            "Encryption: aes256-cbc\nPrivate-Lines: 1\nsynthetic_private_material"
        )
        ssh2_marker = "---- BEGIN SSH2 ENCRYPTED " + "PRIVATE KEY ----"
        bearer_payload = "Authorization: Bearer " + value
        aws_payload = "aws" + "_secret_access_key=" + value

        for archive_payload in (
            tar_with_text("json.txt", json_payload),
            tar_with_text("nested.txt", nested_payload),
            tar_with_text("escaped.txt", escaped_payload),
            tar_with_text("unicode-key.txt", unicode_key_payload),
            tar_with_text("dotted-key.txt", dotted_key_payload),
            tar_with_bytes("utf16.txt", f"{field}={value}".encode("utf-16-le")),
            tar_with_bytes("utf32-le.txt", f"{field}={value}".encode("utf-32-le")),
            tar_with_bytes("utf32-be.txt", f"{field}={value}".encode("utf-32-be")),
            tar_with_text("key.txt", private_marker),
            tar_with_text("encrypted-key.txt", encrypted_marker),
            tar_with_text("dsa-key.txt", dsa_marker),
            tar_with_text("pgp-key.txt", pgp_marker),
            tar_with_text("putty-key.txt", ppk_marker),
            tar_with_text("ssh2-key.txt", ssh2_marker),
            tar_with_text("bearer.txt", bearer_payload),
            tar_with_text("aws.txt", aws_payload),
        ):
            _, errors = HISTORY.inspect_archive(archive_payload)
            self.assertTrue(any("credential" in item for item in errors))
            self.assertNotIn(value, "\n".join(errors))

        for credential_value in (
            json_payload,
            nested_payload,
            escaped_payload,
            unicode_key_payload,
            dotted_key_payload,
            encrypted_marker,
            dsa_marker,
            pgp_marker,
            ppk_marker,
            ssh2_marker,
            bearer_payload,
            aws_payload,
        ):
            state = {
                "schema_version": 1,
                "case_id": credential_value,
                "evidence": [],
                "runs": [],
                "claims": [],
                "documents": [],
                "gates": [],
                "active_index": {},
            }
            state_result = CASE_STATE.validate_state(state)
            self.assertEqual(state_result["errors"][0]["code"], "credential_exposure")
            self.assertNotIn(value, str(state_result))

        for content in (
            nested_payload,
            escaped_payload,
            unicode_key_payload,
            dotted_key_payload,
            encrypted_marker,
            dsa_marker,
            pgp_marker,
            ppk_marker,
            ssh2_marker,
            bearer_payload,
            aws_payload,
        ):
            public_secret, _ = PUBLIC.content_findings(content.encode("utf-8"))
            self.assertTrue(public_secret)
            preview, credential, _ = XLSX.safe_preview(content)
            self.assertTrue(credential)
            self.assertNotIn(value, str(preview))
        self.assertIn(".ppk", PUBLIC.FORBIDDEN_SUFFIXES)
        self.assertEqual(HISTORY.unsafe_path("synthetic.ppk"), "forbidden artifact suffix")
        public_secret, _ = PUBLIC.content_findings(f"{field}={value}".encode("utf-16-le"))
        self.assertTrue(public_secret)
        for encoding in ("utf-32-le", "utf-32-be"):
            public_secret, _ = PUBLIC.content_findings(
                f"{field}={value}".encode(encoding)
            )
            self.assertTrue(public_secret)

        for credential_key in (
            "apiKey",
            "accessToken",
            "refreshToken",
            "clientSecret",
            "privateKey",
            "sonarToken",
            "awsSecretAccessKey",
            "authorization",
        ):
            state = {
                "schema_version": 1,
                "case_id": "CASE-SYNTHETIC",
                "evidence": [],
                "runs": [],
                "claims": [],
                "documents": [],
                "gates": [],
                "active_index": {},
                credential_key: value,
            }
            state_result = CASE_STATE.validate_state(state)
            self.assertEqual(state_result["errors"][0]["code"], "credential_exposure")
            self.assertNotIn(value, str(state_result))

    def test_archive_rejects_absolute_or_traversal_symlink_without_echoing_target(self) -> None:
        target = "/" + "home" + "/synthetic-user/private"
        archived, errors = HISTORY.inspect_archive(tar_with_symlink("safe-link", target))
        self.assertEqual(archived, ["safe-link"])
        self.assertTrue(any("unsafe archive link target" in item for item in errors))
        self.assertNotIn(target, "\n".join(errors))

    def test_history_tree_rejects_deleted_traversal_symlink_without_echo(self) -> None:
        target = "../" + "../outside"
        tree = b"120000 blob abc123\tdeleted-link\0"
        inspected, errors = HISTORY.inspect_historical_symlink_entries(
            tree,
            lambda _oid: target.encode("utf-8"),
        )
        self.assertEqual(inspected, 1)
        self.assertTrue(any("unsafe historical link target" in item for item in errors))
        self.assertNotIn(target, "\n".join(errors))

    def test_history_diagnostics_redact_credential_like_paths(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "path_1234"
        path = f"{field}={marker}"
        _, archive_errors = HISTORY.inspect_archive(
            tar_with_symlink(path, "../../outside")
        )
        tree = b"120000 blob abc123\t" + path.encode("utf-8") + b"\0"
        _, historical_errors = HISTORY.inspect_historical_symlink_entries(
            tree,
            lambda _oid: b"../../outside",
        )
        rendered = "\n".join(archive_errors + historical_errors)
        self.assertIn("[REDACTED_PATH]", rendered)
        self.assertNotIn(marker, rendered)
        self.assertEqual(HISTORY.unsafe_path(path), "credential-like path")
        self.assertEqual(PUBLIC.safe_path_label(path), "[REDACTED_PATH]")

    def test_case_paths_and_machine_paths_are_case_insensitive(self) -> None:
        for path in ("cases/customer.json", "Cases/customer.json", "CASES/customer.json"):
            self.assertEqual(HISTORY.unsafe_path(path), "case data")
            self.assertTrue(PUBLIC.is_root_case_path(path))
        self.assertIsNone(HISTORY.unsafe_path("evals/Cases/synthetic.json"))
        self.assertFalse(PUBLIC.is_root_case_path("evals/Cases/synthetic.json"))

        users_upper = "US" + "ERS"
        home_upper = "HO" + "ME"
        for machine_path in (
            "C:\\" + users_upper + "\\synthetic-user\\private",
            "c:/" + users_upper.lower() + "/synthetic-user/private",
            "/" + users_upper + "/synthetic-user/private",
            "/" + home_upper + "/synthetic-user/private",
            "file:" + "///" + home_upper.lower() + "/synthetic-user/private",
            "file:" + "///C:/" + users_upper.title() + "/synthetic-user/private",
        ):
            public_secret, public_machine = PUBLIC.content_findings(
                machine_path.encode("utf-8")
            )
            history_secret, history_machine = HISTORY.text_findings(machine_path)
            preview, xlsx_secret, xlsx_machine = XLSX.safe_preview(machine_path)
            self.assertFalse(public_secret)
            self.assertTrue(public_machine)
            self.assertFalse(history_secret)
            self.assertTrue(history_machine)
            self.assertFalse(xlsx_secret)
            self.assertTrue(xlsx_machine)
            self.assertEqual(preview, "[REDACTED_SENSITIVE_VALUE]")

        for public_url in (
            "https://example.com/" + users_upper.lower() + "/alice/profile",
            "https://example.com/" + home_upper.lower() + "/alice/profile",
            "https://example.com/C:/" + users_upper.title() + "/alice/profile",
        ):
            self.assertEqual(PUBLIC.text_findings(public_url), (False, False))
            self.assertEqual(HISTORY.text_findings(public_url), (False, False))
            preview, secret, machine_path = XLSX.safe_preview(public_url)
            self.assertFalse(secret)
            self.assertFalse(machine_path)
            self.assertEqual(preview, public_url)

    def test_archive_scan_accepts_sanitized_content(self) -> None:
        archived, errors = HISTORY.inspect_archive(
            tar_with_text("report.txt", "sanitized synthetic evidence")
        )
        self.assertEqual(archived, ["report.txt"])
        self.assertEqual(errors, [])

    def test_zip_and_office_containers_fail_closed_without_unpacking(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "compressed_1234"
        payload = zip_with_text("xl/sharedStrings.xml", f"{field}={marker}")
        self.assertTrue(PUBLIC.is_forbidden_container(payload))
        self.assertIn(".xlsx", PUBLIC.FORBIDDEN_SUFFIXES)
        self.assertEqual(HISTORY.unsafe_path("artifact.xlsx"), "forbidden artifact suffix")

        _, errors = HISTORY.inspect_archive(tar_with_bytes("artifact.bin", payload))
        rendered = "\n".join(errors)
        self.assertIn("nested ZIP/Office container", rendered)
        self.assertNotIn(marker, rendered)

        compound_payload = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64
        self.assertTrue(PUBLIC.is_forbidden_container(compound_payload))
        secret, machine_path, container = HISTORY.stream_text_findings(
            io.BytesIO(compound_payload)
        )
        self.assertFalse(secret)
        self.assertFalse(machine_path)
        self.assertTrue(container)
        _, errors = HISTORY.inspect_archive(
            tar_with_bytes("renamed-compound.txt", compound_payload)
        )
        self.assertIn("nested ZIP/Office container", "\n".join(errors))

        secret, machine_path, container = HISTORY.stream_text_findings(io.BytesIO(payload))
        self.assertFalse(secret)
        self.assertFalse(machine_path)
        self.assertTrue(container)

        prefixed_payload = b"MZ" + b"X" * 126 + payload
        self.assertTrue(zipfile.is_zipfile(io.BytesIO(prefixed_payload)))
        self.assertTrue(PUBLIC.is_forbidden_container(prefixed_payload))
        secret, machine_path, container = HISTORY.stream_text_findings(
            io.BytesIO(prefixed_payload)
        )
        self.assertFalse(secret)
        self.assertFalse(machine_path)
        self.assertTrue(container)
        _, errors = HISTORY.inspect_archive(
            tar_with_bytes("renamed.txt", prefixed_payload)
        )
        self.assertTrue(any("nested ZIP/Office container" in item for item in errors))
        self.assertNotIn(marker, "\n".join(errors))

    def test_xlsx_preview_redacts_credential_like_cell_and_fails(self) -> None:
        field = "api_" + "key"
        value = "synthetic_" + "xlsx_marker_1234"
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "sensitive.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Value"])
            sheet.append(["setting", f"{field}={value}"])
            workbook.save(path)
            result = XLSX.profile(
                path,
                view_kind="rows",
                view_evidence_id="E-VIEW-SAFE",
                view_sheet_index=1,
            )
        self.assertEqual(result["status"], "FAIL")
        self.assertTrue(result["credential_exposure"])
        self.assertFalse(result["rows_present"])
        self.assertNotIn(value, str(result))

    def test_xlsx_visual_gate_is_bound_to_one_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "mixed.xlsx"
            workbook = Workbook()
            confirmed = workbook.active
            confirmed.title = "ConfirmedRows"
            confirmed.append(["Item", "Amount"])
            confirmed.append(["A", 1])
            unconfirmed = workbook.create_sheet("UnconfirmedPropertyTree")
            unconfirmed.append(["Property", "Name"])
            unconfirmed.append(["Type", "String"])
            workbook.save(path)

            result = XLSX.profile(
                path,
                view_kind="rows",
                view_evidence_id="E-VIEW-SHEET-1",
                view_sheet_index=1,
            )

        self.assertTrue(result["rows_present"])
        self.assertEqual(result["view_sheet_index"], 1)
        self.assertTrue(result["sheets"][0]["visual_gate_applies"])
        self.assertTrue(result["sheets"][0]["rows_present"])
        self.assertFalse(result["sheets"][1]["visual_gate_applies"])
        self.assertFalse(result["sheets"][1]["rows_present"])
        self.assertEqual(result["sheets"][1]["view_status"], "unverified/incomplete")

    def test_xlsx_rejects_sensitive_visual_evidence_identifier_without_echo(self) -> None:
        marker = "synthetic_" + "evidence_marker_1234"
        credential_field = "pass" + "word"
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "safe.xlsx"
            workbook = Workbook()
            workbook.active.append(["Name", "Value"])
            workbook.active.append(["A", 1])
            workbook.save(path)
            for evidence_id in (
                f'{{"{credential_field}":"{marker}"}}',
                f"{credential_field}:{marker}",
            ):
                with self.assertRaises(ValueError) as raised:
                    XLSX.profile(
                        path,
                        view_kind="rows",
                        view_evidence_id=evidence_id,
                        view_sheet_index=1,
                    )
                self.assertNotIn(marker, str(raised.exception))

    def test_xlsx_sensitive_filename_uses_fixed_output_name_without_echo(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "filename_1234"
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            path = work / f"{field}={marker}.xlsx"
            workbook = Workbook()
            workbook.active.append(["Name", "Value"])
            workbook.active.append(["A", 1])
            workbook.save(path)
            stdout = io.StringIO()
            with redirect_stdout(stdout):
                status = XLSX.main(
                    [
                        str(path),
                        "--view-kind",
                        "rows",
                        "--view-evidence-id",
                        "E-VIEW-1",
                        "--view-sheet-index",
                        "1",
                    ]
                )
            outputs = list(work.glob("redacted-xlsx-profile-*.json"))
            self.assertEqual(len(outputs), 1)
            output = outputs[0]
            result = json.loads(output.read_text(encoding="utf-8"))
        self.assertEqual(status, 1)
        self.assertEqual(result["status"], "FAIL")
        self.assertNotIn(marker, stdout.getvalue())
        self.assertNotIn(marker, str(result))

    def test_redacted_xlsx_output_names_are_collision_resistant(self) -> None:
        field = "pass" + "word"
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            paths = []
            for index in range(4):
                path = work / f"{field}=synthetic_{index}.xlsx"
                workbook = Workbook()
                workbook.active.append(["Name", "Value"])
                workbook.active.append(["A", index])
                workbook.save(path)
                paths.append(path)

            for path in paths[:2]:
                self.assertEqual(XLSX.main([str(path)]), 1)
            self.assertEqual(
                len(list(work.glob("redacted-xlsx-profile-*.json"))),
                2,
            )

            previous = Path.cwd()
            try:
                os.chdir(work)
                self.assertEqual(COMPARE_XLSX.main(str(paths[0]), str(paths[1])), 1)
                self.assertEqual(COMPARE_XLSX.main(str(paths[2]), str(paths[3])), 1)
            finally:
                os.chdir(previous)
            self.assertEqual(
                len(list(work.glob("redacted-xlsx-comparison-*.diff.json"))),
                2,
            )

    def test_xlsx_cli_failures_do_not_echo_sensitive_names_or_machine_paths(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "missing_1234"
        with tempfile.TemporaryDirectory() as tmp:
            missing = Path(tmp) / f"{field}={marker}.xlsx"
            safe_missing = Path(tmp) / "safe-missing.xlsx"
            for invoke in (
                lambda: XLSX.main([str(missing)]),
                lambda: COMPARE_XLSX.main(str(missing), str(safe_missing)),
            ):
                stdout = io.StringIO()
                stderr = io.StringIO()
                with redirect_stdout(stdout), redirect_stderr(stderr):
                    status = invoke()
                rendered = stdout.getvalue() + stderr.getvalue()
                self.assertEqual(status, 2)
                self.assertNotIn(marker, rendered)
                self.assertNotIn(str(missing), rendered)
                self.assertNotIn("Traceback", rendered)

    def test_xlsx_argument_parser_does_not_echo_sensitive_values(self) -> None:
        marker = "synthetic_" + "parser_secret_1234"
        field = "pass" + "word"
        stdout = io.StringIO()
        stderr = io.StringIO()
        with self.assertRaises(SystemExit) as raised:
            with redirect_stdout(stdout), redirect_stderr(stderr):
                XLSX.main(
                    [
                        "safe.xlsx",
                        "--view-kind",
                        "rows",
                        "--view-evidence-id",
                        "E-1",
                        "--view-sheet-index",
                        f"{field}={marker}",
                    ]
                )
        self.assertEqual(raised.exception.code, 2)
        rendered = stdout.getvalue() + stderr.getvalue()
        self.assertNotIn(marker, rendered)
        self.assertNotIn(field, rendered)

    def test_xlsx_package_metadata_and_comments_fail_closed_without_echo(self) -> None:
        field = "pass" + "word"
        metadata_marker = "synthetic_" + "property_secret_1234"
        comment_marker = "synthetic_" + "comment_secret_1234"
        machine_marker = "C:\\" + "Us" + "ers\\synthetic-user\\private"
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            sensitive = work / "metadata.xlsx"
            safe = work / "safe.xlsx"

            workbook = Workbook()
            workbook.properties.title = f"{field}={metadata_marker}"
            workbook.active["A1"] = "safe"
            workbook.active["A1"].comment = Comment(
                f"{field}={comment_marker}\n{machine_marker}", "synthetic"
            )
            workbook.save(sensitive)

            safe_workbook = Workbook()
            safe_workbook.active["A1"] = "safe"
            safe_workbook.save(safe)

            profile = XLSX.profile(sensitive)
            previous = Path.cwd()
            try:
                os.chdir(work)
                compare_status = COMPARE_XLSX.main(str(sensitive), str(safe))
            finally:
                os.chdir(previous)
            comparison = json.loads(
                (work / "metadata__vs__safe.diff.json").read_text(encoding="utf-8")
            )

        self.assertEqual(profile["status"], "FAIL")
        self.assertTrue(profile["credential_exposure"])
        self.assertTrue(profile["machine_path_exposure"])
        self.assertEqual(compare_status, 1)
        self.assertEqual(comparison["status"], "FAIL")
        rendered = str(profile) + str(comparison)
        self.assertNotIn(metadata_marker, rendered)
        self.assertNotIn(comment_marker, rendered)
        self.assertNotIn(machine_marker, rendered)

    def test_xlsx_binary_vba_member_fails_closed_and_compare_inherits_without_echo(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "vba_secret_1234"
        encoded_secret = f"{field}={marker}".encode("utf-16-le")
        for encoding, width in (
            ("utf-16-le", 2),
            ("utf-16-be", 2),
            ("utf-32-le", 4),
            ("utf-32-be", 4),
        ):
            encoded = f"{field}={marker}".encode(encoding)
            for phase in range(width):
                credential, machine_path = XLSX.stream_security_findings(
                    io.BytesIO((b"x" * phase) + encoded),
                    scan_all_byte_phases=True,
                )
                self.assertTrue(credential)
                self.assertFalse(machine_path)
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            sensitive = work / "macro.xlsx"
            safe = work / "safe.xlsx"

            for path in (sensitive, safe):
                workbook = Workbook()
                workbook.active.append(["Name", "Value"])
                workbook.active.append(["A", 1])
                workbook.save(path)

            with zipfile.ZipFile(sensitive, mode="a") as archive:
                archive.writestr(
                    "xl/vbaProject.bin",
                    encoded_secret,
                )

            with patch.object(
                XLSX,
                "load_workbook",
                wraps=XLSX.load_workbook,
            ) as profile_loader, patch.object(
                COMPARE_XLSX,
                "load_workbook",
                wraps=COMPARE_XLSX.load_workbook,
            ) as compare_loader:
                profile = XLSX.profile(sensitive)
                stdout = io.StringIO()
                stderr = io.StringIO()
                previous = Path.cwd()
                try:
                    os.chdir(work)
                    with redirect_stdout(stdout), redirect_stderr(stderr):
                        compare_status = COMPARE_XLSX.main(str(sensitive), str(safe))
                finally:
                    os.chdir(previous)
            profile_loader.assert_not_called()
            self.assertEqual(compare_loader.call_count, 1)
            comparison = json.loads(
                (work / "macro__vs__safe.diff.json").read_text(encoding="utf-8")
            )

        self.assertEqual(profile["status"], "FAIL")
        self.assertTrue(profile["package_scan_incomplete"])
        self.assertFalse(profile["credential_exposure"])
        self.assertFalse(profile["rows_present"])
        self.assertEqual(profile["view_status"], "package_scan_incomplete")
        self.assertEqual(compare_status, 1)
        self.assertEqual(comparison["status"], "FAIL")
        self.assertTrue(comparison["package_scan_incomplete"])
        rendered = (
            str(profile) + str(comparison) + stdout.getvalue() + stderr.getvalue()
        )
        self.assertNotIn(marker, rendered)

    def test_xlsx_empty_directory_member_is_not_an_unscanned_binary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "directory-entry.xlsx"
            workbook = Workbook()
            workbook.active.append(["Name", "Value"])
            workbook.active.append(["A", 1])
            workbook.save(path)
            with zipfile.ZipFile(path, mode="a") as archive:
                archive.writestr("synthetic-empty-dir/", b"")

            profile = XLSX.profile(path)

        self.assertEqual(profile["status"], "PASS")
        self.assertFalse(profile["package_scan_incomplete"])

    def test_xlsx_directory_named_member_with_payload_fails_closed(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "directory_secret_1234"
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "directory-payload.xlsx"
            workbook = Workbook()
            workbook.active.append(["Name", "Value"])
            workbook.active.append(["A", 1])
            workbook.save(path)
            with zipfile.ZipFile(path, mode="a") as archive:
                archive.writestr(
                    "synthetic-dir/",
                    f"{field}={marker}".encode("utf-8"),
                )

            profile = XLSX.profile(path)

        self.assertEqual(profile["status"], "FAIL")
        self.assertTrue(profile["package_scan_incomplete"])
        self.assertNotIn(marker, str(profile))

    def test_xlsx_zip_comments_and_extra_fields_are_scanned_without_echo(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "zip_metadata_secret_1234"
        secret = f"{field}={marker}".encode("utf-16-le")
        profiles = []
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            for mode in ("archive-comment", "entry-comment", "entry-extra"):
                path = work / f"{mode}.xlsx"
                workbook = Workbook()
                workbook.active.append(["Name", "Value"])
                workbook.active.append(["A", 1])
                workbook.save(path)
                with zipfile.ZipFile(path, mode="a") as archive:
                    if mode == "archive-comment":
                        archive.comment = secret
                    else:
                        info = zipfile.ZipInfo(f"synthetic-{mode}.txt")
                        if mode == "entry-comment":
                            info.comment = secret
                        else:
                            info.extra = (
                                (0x9999).to_bytes(2, "little")
                                + len(secret).to_bytes(2, "little")
                                + secret
                            )
                        archive.writestr(info, b"safe")
                profiles.append(XLSX.profile(path))

            malformed = work / "malformed-extra.xlsx"
            workbook = Workbook()
            workbook.active.append(["Name", "Value"])
            workbook.active.append(["A", 1])
            workbook.save(malformed)
            with zipfile.ZipFile(malformed, mode="a") as archive:
                info = zipfile.ZipInfo("synthetic-malformed-extra.txt")
                info.extra = b"\x99\x99\x05\x00abc"
                archive.writestr(info, b"safe")
            malformed_profile = XLSX.profile(malformed)

        for profile in profiles:
            self.assertEqual(profile["status"], "FAIL")
            self.assertTrue(profile["credential_exposure"])
            self.assertNotIn(marker, str(profile))
        self.assertEqual(malformed_profile["status"], "FAIL")
        self.assertTrue(malformed_profile["package_scan_incomplete"])

    def test_xlsx_local_header_extra_cannot_hide_behind_safe_central_extra(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "local_extra_secret_1234"
        secret = f"{field}={marker}".encode("utf-8")
        filename = b"synthetic-local-extra.txt"
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "local-extra.xlsx"
            workbook = Workbook()
            workbook.active.append(["Name", "Value"])
            workbook.active.append(["A", 1])
            workbook.save(path)
            with zipfile.ZipFile(path, mode="a") as archive:
                info = zipfile.ZipInfo(filename.decode("ascii"))
                info.extra = (
                    (0x9999).to_bytes(2, "little")
                    + len(secret).to_bytes(2, "little")
                    + secret
                )
                archive.writestr(info, b"safe")

            package = bytearray(path.read_bytes())
            offset = 0
            replaced = False
            signature = b"PK\x01\x02"
            while True:
                offset = package.find(signature, offset)
                if offset < 0:
                    break
                name_length = int.from_bytes(package[offset + 28 : offset + 30], "little")
                extra_length = int.from_bytes(package[offset + 30 : offset + 32], "little")
                comment_length = int.from_bytes(package[offset + 32 : offset + 34], "little")
                name_start = offset + 46
                extra_start = name_start + name_length
                if bytes(package[name_start:extra_start]) == filename:
                    safe_extra = (
                        (0x9999).to_bytes(2, "little")
                        + len(secret).to_bytes(2, "little")
                        + b"x" * len(secret)
                    )
                    self.assertEqual(len(safe_extra), extra_length)
                    package[extra_start : extra_start + extra_length] = safe_extra
                    replaced = True
                    break
                offset = extra_start + extra_length + comment_length
            self.assertTrue(replaced)
            path.write_bytes(package)

            with zipfile.ZipFile(path) as archive:
                central_extra = archive.getinfo(filename.decode("ascii")).extra
                self.assertNotIn(secret, central_extra)
            profile = XLSX.profile(path)

        self.assertEqual(profile["status"], "FAIL")
        self.assertTrue(profile["credential_exposure"])
        self.assertNotIn(marker, str(profile))

    def test_xlsx_raw_central_filename_is_scanned_before_decoding(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "central_name_secret_1234"
        encoded_secret = f"{field}={marker}".encode("utf-16-le")
        safe_name = ("x" * (len(encoded_secret) - 4)) + ".bin"
        safe_name_bytes = safe_name.encode("ascii")
        self.assertEqual(len(safe_name_bytes), len(encoded_secret))

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "central-name.xlsx"
            workbook = Workbook()
            workbook.active.append(["Name", "Value"])
            workbook.active.append(["A", 1])
            workbook.save(path)
            with zipfile.ZipFile(path, mode="a") as archive:
                archive.writestr(safe_name, b"safe")

            package = bytearray(path.read_bytes())
            offset = 0
            replaced = False
            signature = b"PK\x01\x02"
            while True:
                offset = package.find(signature, offset)
                if offset < 0:
                    break
                name_length = int.from_bytes(package[offset + 28 : offset + 30], "little")
                extra_length = int.from_bytes(package[offset + 30 : offset + 32], "little")
                comment_length = int.from_bytes(package[offset + 32 : offset + 34], "little")
                name_start = offset + 46
                name_end = name_start + name_length
                if bytes(package[name_start:name_end]) == safe_name_bytes:
                    package[name_start:name_end] = encoded_secret
                    replaced = True
                    break
                offset = name_end + extra_length + comment_length
            self.assertTrue(replaced)
            path.write_bytes(package)

            profile = XLSX.profile(path)

        self.assertEqual(profile["status"], "FAIL")
        self.assertTrue(profile["credential_exposure"])
        self.assertNotIn(marker, str(profile))

    def test_xlsx_lazy_sheet_parse_failure_is_structured_and_fail_closed(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "truncated-sheet.xlsx"
            workbook = Workbook()
            workbook.active.append(["Name", "Value"])
            workbook.active.append(["A", 1])
            workbook.save(path)

            rewritten = io.BytesIO()
            with zipfile.ZipFile(io.BytesIO(path.read_bytes())) as source:
                with zipfile.ZipFile(rewritten, mode="w") as target:
                    for info in source.infolist():
                        content = source.read(info.filename)
                        if info.filename == "xl/worksheets/sheet1.xml":
                            content = b'<worksheet xmlns="http://schemas.openxmlformats.org/'
                        target.writestr(info, content)
            path.write_bytes(rewritten.getvalue())

            profile = XLSX.profile(path)
            snapshot = COMPARE_XLSX.snapshot(path)

        self.assertEqual(profile["status"], "FAIL")
        self.assertTrue(profile["package_scan_incomplete"])
        self.assertFalse(profile["rows_present"])
        data, credential, machine_path, incomplete, _ = snapshot
        self.assertEqual(data, {})
        self.assertFalse(credential)
        self.assertFalse(machine_path)
        self.assertTrue(incomplete)

    def test_xlsx_profile_and_compare_redact_sensitive_sheet_and_cell(self) -> None:
        field = "pass" + "word"
        marker = "synthetic_" + "xlsx_1234"
        with tempfile.TemporaryDirectory() as tmp:
            work = Path(tmp)
            left = work / "left.xlsx"
            right = work / "right.xlsx"
            for path, value in ((left, 1), (right, f"{field}={marker}")):
                workbook = Workbook()
                workbook.active.title = f"{field}={marker}" if path == left else "Data"
                workbook.active.append(["Name", "Value"])
                workbook.active.append(["A", value])
                workbook.save(path)

            profile = XLSX.profile(
                left,
                view_kind="rows",
                view_evidence_id="E-VIEW-1",
                view_sheet_index=1,
            )
            self.assertEqual(profile["status"], "FAIL")
            self.assertNotIn(marker, str(profile))

            previous = Path.cwd()
            try:
                os.chdir(work)
                status = COMPARE_XLSX.main(str(left), str(right))
            finally:
                os.chdir(previous)
            comparison = json.loads(
                (work / "left__vs__right.diff.json").read_text(encoding="utf-8")
            )
            self.assertEqual(status, 1)
            self.assertEqual(comparison["status"], "FAIL")
            self.assertNotIn(marker, str(comparison))


if __name__ == "__main__":
    unittest.main()
